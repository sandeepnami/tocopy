import sys
import os
import shutil
import pandas as pd
from openpyxl import load_workbook
import re
from datetime import datetime, timedelta


#Initialisation
# Define a list of illegal characters for Excel
illegal_chars = ['','→','\x1A','']
# Create a regex pattern to match any of the illegal characters
illegal_chars_pattern = re.compile(f"[{''.join(re.escape(char) for char in illegal_chars)}]")
source_drive = r"\\cvgs070129\esplog\esp\UWdata\sip"
dest_dir = r"C:\Users\ny4007991\OneDrive - Munich Re\Professional\Munich Re\Email Supp\File Processing\202407 July\RE EB file loaded ID1 - Eastern Mutual Insurance Company"
dest_folder = r'inforce'
file_name = r"EQB2363LOC.txt"
mid_dir_path = r"rps\rp0311"
delimitter = "~"
dirs = []
filter_list = ['backup2024'] #['backup2021','backup2022','backup2023','backup2024']
filtered_dirs = []

def contains_illegal_chars(value):
    """
    Check if the value contains any illegal characters.
    """
    if isinstance(value, str):
        return bool(illegal_chars_pattern.search(value))
    return False

def ls(path):
    """
    List all files and directories in the specified path.
    """
    for dir_name in os.listdir(path):
        dir_path = os.path.join(path, dir_name)
        # Get the modification time of the directory
        mod_time = os.path.getmtime(dir_path)
        # Get the Creation time of the directory
        crt_time = os.path.getctime(dir_path)
        
        time_temp = mod_time
        # Convert the time to a datetime object
        mod_time_dt = datetime.fromtimestamp(time_temp)
        fmt_date = mod_time_dt.strftime('%d-%b')
        print(f"{fmt_date}: {dir_name}")
        
def sanitize_data(value):
    """
    Remove illegal characters from the value.
    """
    if isinstance(value, str):
        # Define a regex pattern to match illegal characters
        # Replace illegal characters with an empty string
        return illegal_chars_pattern.sub('', value)
    return value

def search_and_copy(source_drive,dest_folder=''):
    global dirs, filtered_dirs
    entries = os.listdir(source_drive)
    for entry in entries:
        if os.path.isdir(os.path.join(source_drive, entry)):
            dirs.append(entry)
            # Need to be case insensitive for the filter list
            if any(name in entry.lower() for name in filter_list):
                filtered_dirs.append(entry)
    for dir in filtered_dirs:
        dir_path = os.path.join(source_drive, dir,mid_dir_path)
        for root, dirs, files in os.walk(dir_path):
            name_of_file = file_name.split('.')[0]
            matching_files = [file for file in files if name_of_file.lower() in file.lower()]
            # print(f"Matching files in {dir}: {matching_files}")
            for matching_file in matching_files:
                # Construct the full file path
                file_path = os.path.join(root, matching_file)
                new_filename = f"bak{dir[-8:]}_{matching_file}"
                dest_file_path = os.path.join(dest_dir, dest_folder, new_filename)
                # Copy the file to the destination directory
                shutil.copy(file_path, dest_file_path)
                print(f"File {file_name} found in {dir} and copied to {dest_file_path}")

def rename_files(dest_dir,dest_folder=''):
    dest_dir = os.path.join(dest_dir, dest_folder)
    for root, dirs, files in os.walk(dest_dir):
        for file in files:
            try:
                # some criteria to filter only desired files to rename
                if file.startswith('bak') and file.endswith('.xls'):
                    file_name_parts = file.split('-')
                    new_filename = f"{file_name_parts[1][:-4]}_{file_name_parts[0]}.xls" 
                    os.rename(os.path.join(root, file), os.path.join(root, new_filename))
                    print(f"File {file} renamed to {new_filename}")
            except Exception as e:
                print(f"Error renaming file {file}: {e}")

def rename_files_regex(source_dir,pattern, type=''):
    for root, dirs, files in os.walk(source_dir):
        for file in files:
            try:
                # some criteria to filter only desired files to rename like .xlsx
                if file.endswith(type):
                    new_filename = re.sub(pattern, "_", file) 
                    os.rename(os.path.join(root, file), os.path.join(root, new_filename))
                    print(f"File {file} renamed to {new_filename}")
            except Exception as e:
                print(f"Error renaming file {file}: {e}")

def format_excel(excel_path):
    # Load the created Excel file
    print(f"Formatting in excel: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb.active

    # Add filters to the first row
    ws.auto_filter.ref = ws.dimensions
    # Freeze the first row
    ws.freeze_panes = ws['A2']
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the modified Excel file
    wb.save(excel_path)
    print(f"Formatting in excel successful: added Filter to {excel_path}")

#Convert all text files to excel .xlsx format 
def convert_files(dest_dir,dest_folder='',iscontrol=False):
    # print("Inside the convert_files function")
    global delimitter
    dir_path = os.path.join(dest_dir, dest_folder)
    # print("dir_path",dir_path)
    for root, dirs, files in os.walk(dir_path):
        for file in files:
            # print(f"files are {file}")
            if file.lower().endswith('.txt'): # To make sure to inclue .TXT as well
                print(f"Processing.... {file}")
                txtcsv_path = dir_path + "\\" + file
                try:
                    excel_df = pd.read_csv(txtcsv_path, dtype=str,sep=delimitter)
                    # print("Before sanitise",excel_df.columns)
                    excel_df.columns = excel_df.columns.map(sanitize_data)#Just incase illegal characters in column names
                    # Sanitize the DataFrame
                    excel_df = excel_df.map(sanitize_data)
                    # print("after sanitise",excel_df.columns)
                    # excel_file_name = file.lower().replace('.txt', '.xlsx')
                    # $ is the anchor which matches for the end of the string
                    excel_file_name = re.sub(r'\.txt$', '.xlsx', file, flags=re.IGNORECASE)
                    # Check if 'control' is present in the DataFrame
                    if iscontrol:
                        control_record_df = excel_df.tail(1)
                        control_rec_values = control_record_df.iloc[0].astype(str).values
                        control_rec = "~".join(control_rec_values)
                        print("control_rec: ", control_rec)
                        if 'control' in control_rec.lower():    
                            # Assuming you want the 5th column values where 'control' is present in any part of the DataFrame
                            control_date = control_record_df.iloc[0, 4]  # Access the first row and 5th column
                            # print(control_date)
                            excel_file_name = control_date + "_" + excel_file_name
                            excel_path = dir_path + "\\" + excel_file_name    
                        else:
                            print(f"Warning!!! String 'control' not found in the {file}")
                    excel_path = dir_path + "\\" + excel_file_name 
                    excel_df.to_excel(excel_path, index=False)
                    print(f"Conversion to excel successful {file} to {excel_file_name}")
                    format_excel(excel_path)
                except pd.errors.ParserError as e:
                    print(f"Error processing file {file}: {e}")

                
# search_and_copy(source_drive)
# rename_files(dest_dir,dest_folder='excel')

dest_dir = r"C:\Users\ny4007991\OneDrive - Munich Re\Professional\Munich Re\Email Supp\Excel Dumps & Files"
# convert_files(dest_dir)
ls(dest_dir)


# pattern =r"_bak[^_]+_"
# source_dir = os.path.join(dest_dir)
# rename_files_regex(source_dir,pattern=pattern)
